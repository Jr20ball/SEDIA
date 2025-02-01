from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import logging
from concurrent.futures import ThreadPoolExecutor

# Configurar logging
logging.basicConfig(filename="scraping_log.txt", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Configurar Selenium con ChromeDriver
def configure_browser(driver_path):
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Modo sin interfaz gráfica
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--log-level=3')  # Reduce logs de Chrome
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36")
    
    service = Service(driver_path)
    return webdriver.Chrome(service=service, options=chrome_options)

# Función para extraer datos
def extract_subsidy_data(url, driver_path, index, total):
    print(f"Procesando {index+1}/{total}: {url}")
    browser = configure_browser(driver_path)
    data = {"URL": url}  

    try:
        browser.get(url)
        wait = WebDriverWait(browser, 5)

        try:
            print_div = wait.until(EC.presence_of_element_located((By.ID, "print")))
        except:
            logging.warning(f"No se encontró el contenedor en {url}")
            return data  

        campos = print_div.find_elements(By.CLASS_NAME, "titulo-campo")
        for campo in campos:
            titulo = campo.text.strip()
            try:
                valor_divs = campo.find_elements(By.XPATH, "following-sibling::div[contains(@class, 'padding-top')]")
                valor = " ".join([div.text.strip() for div in valor_divs if div.text.strip()]) if valor_divs else "No disponible"
            except:
                valor = "No disponible"

            data[titulo] = valor  

    except Exception as e:
        logging.error(f"Error en {url}: {e}")
    finally:
        browser.quit()

    return data

# Función para formatear Excel
def format_excel(output_file):
    wb = load_workbook(output_file)
    ws = wb.active

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    data_fill1 = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    data_fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = data_fill1 if cell.row % 2 == 0 else data_fill2

    wb.save(output_file)
    print(f"🎨 Formato aplicado en {output_file}")

# Procesar múltiples URLs en paralelo
def process_subsidy_urls(driver_path, urls, output_file):
    total_urls = len(urls)
    with ThreadPoolExecutor(max_workers=6) as executor:
        results = executor.map(lambda args: extract_subsidy_data(args[0], driver_path, args[1], total_urls), [(url, i) for i, url in enumerate(urls)])
    
    all_data = list(results)
    df = pd.DataFrame(all_data).fillna("No disponible")  
    df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"✅ Datos exportados a {output_file}")
    format_excel(output_file)

# Cargar las URLs desde el archivo CSV
file_path = "C:\\Users\\albar\\Desktop\\enlaces_infosubvenciones.csv"
df_urls = pd.read_csv(file_path, sep=';')
urls = df_urls["Enlace"].dropna().tolist()
cleaned_urls = [url.replace("https://www.infosubvenciones.eshttps://www.infosubvenciones.es", "https://www.infosubvenciones.es") for url in urls]

# Ruta del driver
driver_path = "C:\\Users\\albar\\Downloads\\chromedriver-win64\\chromedriver.exe"

# Archivo de salida
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
output_file = os.path.join(desktop, "subvenciones_formato.xlsx")

# Ejecutar el proceso
process_subsidy_urls(driver_path, cleaned_urls, output_file)
